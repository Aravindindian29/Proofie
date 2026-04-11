import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings

logger = logging.getLogger(__name__)


class ExcelService:
    @staticmethod
    def create_test_cases_excel(test_cases, risk_areas=None, regression_scope=None, filename='test_cases.xlsx'):
        try:
            wb = Workbook()
            
            # Sheet 1: Test Cases
            ws = wb.active
            ws.title = "Test Cases"
            
            headers = [
                'Test Case ID',
                'Scenario',
                'Steps',
                'Expected Result',
                'Priority',
                'Type'
            ]
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Priority color fills
            priority_fills = {
                'high': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'),
                'medium': PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid'),
                'low': PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
            }
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row_num, test_case in enumerate(test_cases, 2):
                # Column 1: Test Case ID
                ws.cell(row=row_num, column=1).value = test_case.get('test_case_id', f'TC{row_num-1:03d}')
                
                # Column 2: Scenario
                ws.cell(row=row_num, column=2).value = test_case.get('scenario', test_case.get('title', ''))
                
                # Column 3: Steps
                steps = test_case.get('steps', [])
                if isinstance(steps, list):
                    steps_text = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(steps)])
                else:
                    steps_text = str(steps)
                ws.cell(row=row_num, column=3).value = steps_text
                
                # Column 4: Expected Result - ensure it has meaningful content
                expected_result = test_case.get('expected_result', test_case.get('expected', ''))
                if not expected_result or expected_result.strip() == '':
                    # Generate a default expected result based on scenario if missing
                    scenario = test_case.get('scenario', '')
                    if 'error' in scenario.lower() or 'invalid' in scenario.lower():
                        expected_result = 'System displays appropriate error message and prevents invalid action'
                    elif 'success' in scenario.lower() or 'valid' in scenario.lower():
                        expected_result = 'Operation completes successfully with confirmation message'
                    else:
                        expected_result = 'System behaves as expected per requirements'
                ws.cell(row=row_num, column=4).value = expected_result
                
                # Column 5: Priority
                priority = test_case.get('priority', 'medium').lower()
                priority_cell = ws.cell(row=row_num, column=5)
                priority_cell.value = priority.capitalize()
                if priority in priority_fills:
                    priority_cell.fill = priority_fills[priority]
                
                # Column 6: Type
                ws.cell(row=row_num, column=6).value = test_case.get('type', 'functional')
                
                # Apply formatting to all cells
                for col_num in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            column_widths = {
                'A': 15,   # Test Case ID
                'B': 40,   # Scenario
                'C': 60,   # Steps
                'D': 45,   # Expected Result
                'E': 12,   # Priority
                'F': 15    # Type
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            ws.row_dimensions[1].height = 25
            for row in range(2, len(test_cases) + 2):
                ws.row_dimensions[row].height = 60
            
            # Sheet 2: Risk Areas
            if risk_areas:
                ws_risk = wb.create_sheet(title="Risk Areas")
                
                ws_risk.cell(row=1, column=1).value = "Risk Areas"
                ws_risk.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws_risk.merge_cells('A1:B1')
                
                current_row = 3
                
                high_risks = risk_areas.get('high_risk', [])
                if high_risks:
                    ws_risk.cell(row=current_row, column=1).value = "High Risk"
                    ws_risk.cell(row=current_row, column=1).fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    ws_risk.cell(row=current_row, column=1).font = Font(bold=True, color='FFFFFF')
                    current_row += 1
                    for risk in high_risks:
                        ws_risk.cell(row=current_row, column=1).value = "•"
                        ws_risk.cell(row=current_row, column=2).value = risk
                        ws_risk.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
                        current_row += 1
                    current_row += 1
                
                medium_risks = risk_areas.get('medium_risk', [])
                if medium_risks:
                    ws_risk.cell(row=current_row, column=1).value = "Medium Risk"
                    ws_risk.cell(row=current_row, column=1).fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                    ws_risk.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    for risk in medium_risks:
                        ws_risk.cell(row=current_row, column=1).value = "•"
                        ws_risk.cell(row=current_row, column=2).value = risk
                        ws_risk.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
                        current_row += 1
                    current_row += 1
                
                low_risks = risk_areas.get('low_risk', [])
                if low_risks:
                    ws_risk.cell(row=current_row, column=1).value = "Low Risk"
                    ws_risk.cell(row=current_row, column=1).fill = PatternFill(start_color='6BCF7F', end_color='6BCF7F', fill_type='solid')
                    ws_risk.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    for risk in low_risks:
                        ws_risk.cell(row=current_row, column=1).value = "•"
                        ws_risk.cell(row=current_row, column=2).value = risk
                        ws_risk.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
                        current_row += 1
                
                ws_risk.column_dimensions['A'].width = 5
                ws_risk.column_dimensions['B'].width = 80
            
            # Sheet 3: Regression Scope
            if regression_scope:
                ws_scope = wb.create_sheet(title="Regression Scope")
                
                ws_scope.cell(row=1, column=1).value = "Regression Scope"
                ws_scope.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws_scope.merge_cells('A1:B1')
                
                current_row = 3
                
                for category, items in regression_scope.items():
                    category_title = category.replace('_', ' ').title()
                    ws_scope.cell(row=current_row, column=1).value = category_title
                    ws_scope.cell(row=current_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    ws_scope.cell(row=current_row, column=1).font = Font(bold=True, color='FFFFFF')
                    current_row += 1
                    
                    for item in items:
                        ws_scope.cell(row=current_row, column=1).value = "•"
                        ws_scope.cell(row=current_row, column=2).value = item
                        ws_scope.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
                        current_row += 1
                    current_row += 1
                
                ws_scope.column_dimensions['A'].width = 5
                ws_scope.column_dimensions['B'].width = 80
            
            test_cases_dir = os.path.join(settings.MEDIA_ROOT, 'test_cases')
            os.makedirs(test_cases_dir, exist_ok=True)
            
            file_path = os.path.join(test_cases_dir, filename)
            wb.save(file_path)
            
            logger.info(f"Excel file created: {file_path}")
            
            relative_path = os.path.join('test_cases', filename)
            file_url = f"{settings.MEDIA_URL}{relative_path}".replace('\\', '/')
            
            return {
                'file_path': file_path,
                'file_url': file_url,
                'filename': filename
            }
        
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            raise
